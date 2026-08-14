"""
ingestion/parsers/excel_parser.py
───────────────────────────────────
Excel (.xlsx) parser using openpyxl.

Extracts data from all sheets. Each sheet becomes a separate block
with a structured text representation:
  - Sheet name as a header
  - Column headers as the first row
  - All data rows serialized as tab-separated values

Also captures basic chart presence (chart titles if available) so
the LLM knows visual content exists even if it can't see images.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any, Optional

import structlog
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart

logger = structlog.get_logger(__name__)

# Max rows to extract per sheet to avoid absurdly large chunks
MAX_ROWS_PER_SHEET = 5000


@dataclass
class SheetContent:
    sheet_name: str
    headers: list[str]
    rows: list[list[Any]]
    chart_titles: list[str] = field(default_factory=list)
    row_count: int = 0
    col_count: int = 0

    def to_text(self) -> str:
        """Convert sheet content to a readable text representation."""
        lines: list[str] = [f"[Sheet: {self.sheet_name}]"]

        if self.chart_titles:
            lines.append(f"[Charts: {', '.join(self.chart_titles)}]")

        if self.headers:
            lines.append("\t".join(str(h) for h in self.headers))

        for row in self.rows[:MAX_ROWS_PER_SHEET]:
            row_str = "\t".join(
                "" if v is None else str(v) for v in row
            )
            if row_str.strip():
                lines.append(row_str)

        if self.row_count > MAX_ROWS_PER_SHEET:
            lines.append(
                f"[Note: {self.row_count - MAX_ROWS_PER_SHEET} additional rows truncated]"
            )

        return "\n".join(lines)


# Every pre-2007 .xls file begins with this OLE2 compound-document signature.
# Checking the bytes rather than the file extension means a mislabelled upload is
# still diagnosed correctly.
_OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _looks_like_legacy_xls(file_bytes: bytes) -> bool:
    """True if these bytes are an old-format .xls rather than an .xlsx."""
    return file_bytes[:8] == _OLE2_SIGNATURE


def _parse_legacy_xls(file_bytes: bytes, filename: str) -> list[SheetContent]:
    """
    Read a pre-2007 .xls workbook via xlrd.

    Kept behind openpyxl rather than dispatched on the file extension, because
    files are routinely misnamed — a .xls that is really an .xlsx is common
    enough that trusting the extension would reject valid uploads.

    xlrd 2.x reads .xls only (its .xlsx support was removed), which is exactly
    the split we want: openpyxl for the modern package format, xlrd for the
    binary one.
    """
    try:
        import xlrd
    except ImportError as e:
        raise ValueError(
            f"'{filename}' is in the older Excel format (.xls), and the reader for "
            "it is not installed. Please re-save the file as .xlsx and upload again."
        ) from e

    try:
        book = xlrd.open_workbook(file_contents=file_bytes)
    except Exception as e:
        raise ValueError(f"Failed to open Excel file '{filename}': {e}") from e

    def _cell(value: Any, ctype: int, datemode: int) -> Any:
        # xlrd reports dates as floats with a separate type flag; without this
        # every date in the sheet reads as a five-digit number.
        if ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate.xldate_as_datetime(value, datemode).isoformat(" ")
            except Exception:
                return value
        if ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(value)
        if ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
            return int(value)
        return value

    sheets: list[SheetContent] = []
    for ws in book.sheets():
        if ws.nrows == 0:
            continue
        rows = [
            [_cell(ws.cell_value(r, c), ws.cell_type(r, c), book.datemode)
             for c in range(ws.ncols)]
            for r in range(min(ws.nrows, MAX_ROWS_PER_SHEET + 1))
        ]
        headers = [str(v) if v is not None else "" for v in rows[0]]
        sheets.append(
            SheetContent(
                sheet_name=ws.name,
                headers=headers,
                rows=rows[1:],
                chart_titles=[],
                row_count=ws.nrows - 1,
                col_count=ws.ncols,
            )
        )

    logger.info(
        "excel_parser.complete_legacy_xls",
        filename=filename,
        total_sheets=len(sheets),
        total_rows=sum(s.row_count for s in sheets),
    )
    return sheets


def parse_excel(
    file_bytes: bytes,
    filename: str = "document.xlsx",
    data_only: bool = True,
) -> list[SheetContent]:
    """
    Parse an .xlsx file from raw bytes.

    Note this handles the modern .xlsx format only. See _looks_like_legacy_xls.

    Args:
        data_only: If True, reads computed cell values rather than formulas.

    Returns a list of SheetContent, one per worksheet.
    """
    try:
        wb = load_workbook(
            filename=io.BytesIO(file_bytes),
            data_only=data_only,
            read_only=False,  # read_only=True breaks chart access
        )
    except Exception as e:
        # openpyxl reads the modern .xlsx/.xlsm package format only. A pre-2007
        # .xls is an OLE2 binary, which xlrd handles instead.
        if _looks_like_legacy_xls(file_bytes):
            return _parse_legacy_xls(file_bytes, filename)
        raise ValueError(f"Failed to open Excel file '{filename}': {e}") from e

    sheets: list[SheetContent] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip completely empty sheets
        if ws.max_row is None or ws.max_row == 0:
            continue

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # First non-empty row is treated as headers
        headers: list[str] = []
        data_rows: list[list[Any]] = []

        for i, row in enumerate(all_rows):
            if i == 0:
                headers = [str(cell) if cell is not None else "" for cell in row]
            else:
                data_rows.append(list(row))

        # Extract chart titles
        chart_titles: list[str] = []
        for chart in ws._charts:  # type: ignore[attr-defined]
            title = _get_chart_title(chart)
            if title:
                chart_titles.append(title)

        sheet = SheetContent(
            sheet_name=sheet_name,
            headers=headers,
            rows=data_rows,
            chart_titles=chart_titles,
            row_count=len(data_rows),
            col_count=len(headers),
        )
        sheets.append(sheet)

    logger.info(
        "excel_parser.complete",
        filename=filename,
        total_sheets=len(sheets),
        total_rows=sum(s.row_count for s in sheets),
    )
    return sheets


def _get_chart_title(chart: Any) -> Optional[str]:
    """Safely extract a chart's title string."""
    try:
        if hasattr(chart, "title") and chart.title:
            t = chart.title
            if hasattr(t, "tx") and t.tx:
                return str(t.tx)
            return str(t)
    except Exception:
        pass
    return None
